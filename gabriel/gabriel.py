import PySimpleGUI as sg
import openpyxl

# Defina o tema personalizado no início do programa


# Função para adicionar um registro à planilha
def adicionar_registro(nome, valor, quantidade, data):
    try:
        workbook = openpyxl.load_workbook("banca.xlsx")
        worksheet = workbook.active
        linha = [nome, valor, quantidade, data]
        worksheet.append(linha)
        workbook.save("banca.xlsx")
        return True
    except Exception as e:
        return str(e)

# Função para listar os registros da planilha
def listar_registros():
    try:
        workbook = openpyxl.load_workbook("banca.xlsx")
        worksheet = workbook.active
        registros = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return registros
    except Exception as e:
        return str(e)

# Função para remover registros da planilha por índices
def remover_registros(indices):
    try:
        workbook = openpyxl.load_workbook("banca.xlsx")
        worksheet = workbook.active

        # Remove os registros da planilha com base nos índices
        for indice in reversed(indices):  # Remova na ordem inversa para evitar problemas de índice
            worksheet.delete_rows(indice + 2)

        workbook.save("banca.xlsx")
        return True
    except Exception as e:
        return str(e)

# Função para calcular o valor total dos registros
def calcular_valor_total(registros):
    total = 0
    for registro in registros:
        valor, quantidade = registro[1], registro[2]
        if valor is not None and quantidade is not None:
            try:
                valor = float(valor)  # Converte para float
                quantidade = int(quantidade)  # Converte para int
                total += valor * quantidade
            except (ValueError, TypeError):
                print(f"Erro ao converter valores: {valor}, {quantidade}")
    return total

# Função para calcular o valor total dos registros de um mês específico
def calcular_valor_total_mes(registros, mes):
    total = 0
    for registro in registros:
        data = registro[3]
        if data and data.month == mes:
            valor, quantidade = registro[1], registro[2]
            if valor is not None and quantidade is not None:
                try:
                    valor = float(valor)  # Converte para float
                    quantidade = int(quantidade)  # Converte para int
                    total += valor * quantidade
                except (ValueError, TypeError):
                    print(f"Erro ao converter valores: {valor}, {quantidade}")
    return total

# Layout da página principal
layout_principal = [
    [sg.Text("Nome:"), sg.InputText(key="nome")],
    [sg.Text("Valor:"), sg.InputText(key="valor")],
    [sg.Text("Quantidade:"), sg.InputText(key="quantidade")],
    [sg.Text("Data (AAAA-MM-DD):"), sg.InputText(key="data")],
    [sg.Button("Adicionar Registro")],
    [sg.Table(values=[], headings=["Índice", "Nome", "Valor", "Quantidade", "Data"], auto_size_columns=False, num_rows=10, justification='right', key="table")],
    [sg.Text("Índice(s) do Registro para Remover:"), sg.InputText(key="selecionado")],
    [sg.Button("Remover Selecionado"), sg.Button("Listar Registros"), sg.Button("Calcular Valor Total"),sg.Button("Meses"), sg.Button("Sair")],
]

# Cria a janela da interface
window = sg.Window("Cadastro de Registros em Fpdoces", layout_principal)

# Loop principal da aplicação
while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED or event == "Sair":
        break


    if event == "Adicionar Registro":
        nome, valor, quantidade, data = values["nome"], values["valor"], values["quantidade"], values["data"]
        if nome and valor and quantidade and data and adicionar_registro(nome, valor, quantidade, data):
            sg.popup("Registro adicionado com sucesso!")
        else:
            sg.popup_error("Erro ao adicionar registro.")

    if event == "Listar Registros":
        registros = listar_registros()
        if isinstance(registros, list):
            registros_com_indices = [[i] + registro for i, registro in enumerate(registros)]
            window["table"].update(values=registros_com_indices)
        else:
            sg.popup_error("Erro ao listar registros.")

    if event == "Remover Selecionado":
        selecionados = values["selecionado"].split(",")
        indices_para_remover = []
        for selecionado in selecionados:
            try:
                indice = int(selecionado)
                indices_para_remover.append(indice)
            except ValueError:
                sg.popup_error("Índice inválido. Digite números separados por vírgula.")

        if indices_para_remover:
            if remover_registros(indices_para_remover):
                sg.popup("Registros removidos com sucesso!")
            else:
                sg.popup_error("Erro ao remover registros.")

    if event == "Calcular Valor Total":
        registros = listar_registros()
        if isinstance(registros, list):
            total = calcular_valor_total(registros)
            sg.popup(f"Valor Total: R$ {total:.2f}")

    if event == "Meses":
        meses_layout = [
            [sg.Text("Selecione o mês (1 a 12):"), sg.InputText(key="mes_selecionado")],
            [sg.Button("Selecionar Mês")],
        ]

        meses_window = sg.Window("Selecionar Mês", meses_layout)

        while True:
            meses_event, meses_values = meses_window.read()

            if meses_event == sg.WIN_CLOSED:
                break

            if meses_event == "Selecionar Mês":
                mes_selecionado = meses_values["mes_selecionado"]
                try:
                    mes_selecionado = int(mes_selecionado)
                    if 1 <= mes_selecionado <= 12:
                        registros = listar_registros()
                        total_mes = calcular_valor_total_mes(registros, mes_selecionado)
                        sg.popup(f"Valor Total do Mês {mes_selecionado}: R$ {total_mes:.2f}")
                    else:
                        sg.popup_error("Mês inválido. Selecione um mês válido (1 a 12).")
                except (ValueError, TypeError):
                    sg.popup_error("Mês inválido. Selecione um mês válido (1 a 12).")

        meses_window.close()


# Código para abrir a janela de "Meses"
layout_meses = [
    [sg.Text("Selecione um mês para calcular o valor total:")],
    [sg.Combo([str(i) for i in range(1, 13)], key="mes_combo", default_value="1"), sg.Button("Calcular Valor Total do Mês"), sg.Button("Fechar")]
]